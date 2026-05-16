from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import logging
import re
import unicodedata
import uuid

from django.db import connection, OperationalError, ProgrammingError, transaction
from django.db.models import Q
from openpyxl import load_workbook

from apps.core.models import MinutaRomaneio, MinutaRomaneioItem
from apps.nf.models import EntradaNF, NotaFiscal, nota_fiscal_bairro_valor
from apps.nf.services.importador_xml import ImportacaoXMLError, analisar_xml_nfe
from apps.nf.services.xml_storage_service import XMLStorageUnavailableError, open_entrada_xml


logger = logging.getLogger(__name__)


MINUTA_STATUS_NAO_ENCONTRADA = {'NF NÃO LOCALIZADA', 'XML INVALIDO', 'NF COM PROBLEMA'}
MINUTA_STATUS_INCONSISTENTE = {'NF CANCELADA', 'NF DENEGADA', 'NF INCONSISTENTE', 'NF BLOQUEADA', 'NF INATIVA'}


class MinutaImportacaoError(Exception):
    pass


def _diagnostico_tabelas_minuta():
    tabelas_esperadas = {'core_minutaromaneio', 'core_minutaromaneioitem'}
    diagnostico = {
        'schema_detectado': connection.vendor,
        'alias': connection.alias,
        'tabelas_encontradas': [],
        'tabelas_faltantes': [],
        'erro': '',
        'resultado_validacao': False,
    }
    try:
        tabelas = set(connection.introspection.table_names())
    except (OperationalError, ProgrammingError) as exc:
        diagnostico['erro'] = str(exc)
        logger.exception('DEBUG MINUTA: falha ao consultar table_names() da conexao atual.')
        return diagnostico

    diagnostico['tabelas_encontradas'] = sorted(tabelas.intersection(tabelas_esperadas))
    diagnostico['tabelas_faltantes'] = sorted(tabelas_esperadas - tabelas)
    diagnostico['resultado_validacao'] = not diagnostico['tabelas_faltantes']
    logger.info(
        'DEBUG MINUTA: schema_detectado=%s alias=%s tabelas_encontradas=%s tabelas_faltantes=%s validacao=%s',
        diagnostico['schema_detectado'],
        diagnostico['alias'],
        diagnostico['tabelas_encontradas'],
        diagnostico['tabelas_faltantes'],
        diagnostico['resultado_validacao'],
    )
    return diagnostico


def _mensagem_erro_estrutura_minuta(diagnostico):
    if diagnostico.get('erro'):
        return (
            'ERRO REAL MINUTA: falha ao validar estrutura no banco. '
            f"schema={diagnostico.get('schema_detectado')} alias={diagnostico.get('alias')} erro={diagnostico.get('erro')}"
        )
    return (
        'ERRO REAL MINUTA: validação da estrutura retornou falso. '
        f"schema={diagnostico.get('schema_detectado')} alias={diagnostico.get('alias')} "
        f"tabelas_encontradas={diagnostico.get('tabelas_encontradas')} "
        f"tabelas_faltantes={diagnostico.get('tabelas_faltantes')}"
    )


def get_minuta_inconsistencias(linhas):
    nfs_nao_encontradas = 0
    xml_erros = 0
    duplicidades = 0
    inconsistencias = 0

    for linha in linhas:
        status = (linha.get('status') or '').strip()
        duplicado = bool(linha.get('duplicado'))
        tem_nf_vinculada = bool(linha.get('nf_id'))

        if duplicado:
            duplicidades += 1
        if status == 'XML INVALIDO':
            xml_erros += 1
        if status in MINUTA_STATUS_NAO_ENCONTRADA or not tem_nf_vinculada:
            nfs_nao_encontradas += 1
        if status in MINUTA_STATUS_INCONSISTENTE:
            inconsistencias += 1

    return {
        'nfs_nao_encontradas': nfs_nao_encontradas,
        'xml_erros': xml_erros,
        'duplicidades': duplicidades,
        'inconsistencias': inconsistencias,
        'possui_alertas': any([nfs_nao_encontradas, xml_erros, duplicidades, inconsistencias]),
    }


def _texto_limpo(valor):
    if valor is None:
        return ''
    return str(valor).strip()


def _normalizar_cabecalho(valor):
    texto = unicodedata.normalize('NFKD', _texto_limpo(valor)).encode('ascii', 'ignore').decode('ascii')
    texto = texto.lower()
    return re.sub(r'[^a-z0-9]+', '', texto)


def _parse_data(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = _texto_limpo(valor)
    if not texto:
        return None
    for formato in ('%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def _parse_decimal(valor):
    texto = _texto_limpo(valor)
    if not texto:
        return Decimal('0')
    if ',' in texto and '.' in texto:
        if texto.rfind(',') > texto.rfind('.'):
            texto = texto.replace('.', '').replace(',', '.')
        else:
            texto = texto.replace(',', '')
    elif ',' in texto:
        texto = texto.replace(',', '.')
    try:
        return Decimal(texto)
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _parse_int(valor):
    texto = _texto_limpo(valor)
    if not texto:
        return None
    try:
        return int(float(texto.replace(',', '.')))
    except ValueError:
        return None


def _extrair_placa(veiculo):
    texto = _texto_limpo(veiculo)
    if ' - ' in texto:
        texto = texto.split(' - ', 1)[1]
    if '/' in texto:
        texto = texto.split('/', 1)[0]
    return texto.strip().upper()


def _extrair_nome_pessoa(campo):
    texto = _texto_limpo(campo)
    if ' - ' in texto:
        return texto.split(' - ', 1)[1].strip()
    return texto


def _linha_tem_conteudo(linha):
    return any(_texto_limpo(valor) for valor in linha)


def _localizar_linha_cabecalho(rows, cabecalhos):
    esperados = {_normalizar_cabecalho(valor) for valor in cabecalhos}
    for index, row in enumerate(rows[:20]):
        normalizados = {_normalizar_cabecalho(valor) for valor in row if _texto_limpo(valor)}
        if esperados.issubset(normalizados):
            return index
    raise MinutaImportacaoError('Não foi possível localizar o cabeçalho da planilha de romaneio.')


def _mapear_indices_cabecalho(row, obrigatorios=None):
    indices = {}
    for index, valor in enumerate(row):
        chave = _normalizar_cabecalho(valor)
        if chave and chave not in indices:
            indices[chave] = index
    faltantes = set(obrigatorios or []) - set(indices)
    if faltantes:
        raise MinutaImportacaoError('A planilha não possui as colunas mínimas para importação da minuta.')
    return indices


def _obter_valor(linha, indices, chave):
    index = indices.get(chave)
    if index is None or index >= len(linha):
        return ''
    return linha[index]


def _serializar_data(valor):
    if isinstance(valor, (datetime, date)):
        return valor.isoformat()
    return valor or ''


def _bairro_nf(nf):
    if nf is None:
        return ''
    return (nota_fiscal_bairro_valor(nf) or getattr(getattr(nf, 'rota', None), 'bairro', '') or '').strip()


def _status_operacional_nf_minuta(nf):
    if nf.status == NotaFiscal.Status.PENDENTE:
        return 'XML IMPORTADO'
    if nf.status == NotaFiscal.Status.EM_CONFERENCIA:
        return 'EM CONFERENCIA'
    if nf.status == NotaFiscal.Status.CONCLUIDO:
        return 'FINALIZADA'
    if nf.status == NotaFiscal.Status.CONCLUIDO_COM_RESTRICAO:
        return 'FINALIZADA COM RESTRICAO'
    if nf.status == NotaFiscal.Status.NORMAL:
        return 'XML IMPORTADO'
    if nf.status == NotaFiscal.Status.LIBERADA_COM_RESTRICAO:
        return 'LIBERADA COM RESTRICAO'
    return nf.status.replace('_', ' ')


def _nf_importavel_para_minuta(nf):
    if nf is None:
        return False
    if nf.status_fiscal != NotaFiscal.StatusFiscal.AUTORIZADA:
        return False
    if not nf.ativa or nf.bloqueada:
        return False
    if nf.status in {NotaFiscal.Status.BLOQUEADA_COM_RESTRICAO, NotaFiscal.Status.INCONSISTENTE}:
        return False
    return True


def _carregar_xml_entrada_por_numero(entrada_por_numero, usuario):
    xml_por_numero = {}
    for numero_nota, entrada in entrada_por_numero.items():
        if entrada is None or entrada.status == EntradaNF.Status.PROCESSADO:
            continue
        try:
            with open_entrada_xml(entrada, user=usuario) as arquivo_xml:
                documento = analisar_xml_nfe(arquivo_xml)
        except (ImportacaoXMLError, XMLStorageUnavailableError) as exc:
            xml_por_numero[numero_nota] = {
                'cliente_nome': '',
                'bairro': '',
                'status_fiscal': None,
                'xml_invalido': True,
                'motivo': f'XML inválido ou indisponível para a NF {numero_nota}: {exc}',
            }
            continue
        xml_por_numero[numero_nota] = {
            'cliente_nome': (documento.cliente_nome or '').strip(),
            'bairro': (documento.bairro or '').strip(),
            'status_fiscal': documento.status_fiscal,
            'xml_invalido': False,
            'motivo': '',
        }
    return xml_por_numero


def _resolver_status_minuta_nf(nf, entrada, xml_info=None):
    xml_info = xml_info or {}
    if _nf_importavel_para_minuta(nf):
        return {
            'status': _status_operacional_nf_minuta(nf),
            'importavel': True,
            'motivo': '',
            'nf': nf,
        }

    if nf is not None:
        if nf.status_fiscal == NotaFiscal.StatusFiscal.CANCELADA:
            return {
                'status': 'NF CANCELADA',
                'importavel': False,
                'motivo': 'NF cancelada não pode ser usada na minuta.',
                'nf': None,
            }
        if nf.status_fiscal == NotaFiscal.StatusFiscal.DENEGADA:
            return {
                'status': 'NF DENEGADA',
                'importavel': False,
                'motivo': 'NF denegada não pode ser usada na minuta.',
                'nf': None,
            }
        if nf.status == NotaFiscal.Status.INCONSISTENTE:
            return {
                'status': 'NF INCONSISTENTE',
                'importavel': False,
                'motivo': 'NF inconsistente não pode ser usada na minuta.',
                'nf': None,
            }
        if nf.bloqueada or nf.status == NotaFiscal.Status.BLOQUEADA_COM_RESTRICAO:
            return {
                'status': 'NF BLOQUEADA',
                'importavel': False,
                'motivo': 'NF bloqueada/restrita não pode ser usada na minuta.',
                'nf': None,
            }
        if not nf.ativa:
            return {
                'status': 'NF INATIVA',
                'importavel': False,
                'motivo': 'NF inativa não pode ser usada na minuta.',
                'nf': None,
            }

    if entrada is not None:
        if xml_info.get('xml_invalido'):
            return {
                'status': 'XML INVALIDO',
                'importavel': False,
                'motivo': xml_info.get('motivo') or 'XML inválido ou indisponível para vinculação da NF.',
                'nf': None,
            }
        if xml_info.get('status_fiscal') == NotaFiscal.StatusFiscal.CANCELADA:
            return {
                'status': 'NF CANCELADA',
                'importavel': False,
                'motivo': 'NF cancelada não pode ser usada na minuta.',
                'nf': None,
            }
        if xml_info.get('status_fiscal') == NotaFiscal.StatusFiscal.DENEGADA:
            return {
                'status': 'NF DENEGADA',
                'importavel': False,
                'motivo': 'NF denegada não pode ser usada na minuta.',
                'nf': None,
            }
        if entrada.status == EntradaNF.Status.AGUARDANDO:
            return {
                'status': 'AGUARDANDO LIBERACAO',
                'importavel': True,
                'motivo': '',
                'nf': None,
            }
        if entrada.status == EntradaNF.Status.PROCESSADO:
            return {
                'status': 'NF COM PROBLEMA',
                'importavel': False,
                'motivo': 'NF possui problema no pré-processamento e não pode entrar na minuta.',
                'nf': None,
            }
        if entrada.status == EntradaNF.Status.LIBERADO:
            return {
                'status': 'LIBERADA',
                'importavel': True,
                'motivo': '',
                'nf': None,
            }

    return {
        'status': 'NF NÃO LOCALIZADA',
        'importavel': False,
        'motivo': 'NF não localizada no WMS e sem entrada liberada válida.',
        'nf': None,
    }


def _validar_preview_minuta_confirmavel(preview):
    linhas_bloqueadas = [linha for linha in preview.get('linhas', []) if not linha.get('importavel')]
    if not linhas_bloqueadas:
        return
    exemplos = ', '.join(f"{linha['numero_nota']} ({linha['status']})" for linha in linhas_bloqueadas[:5])
    raise MinutaImportacaoError(
        'A importação da minuta foi bloqueada. Somente NFs válidas, sem cancelamento, bloqueio, inconsistência ou XML inválido podem ser utilizadas. '
        f'Pendências encontradas: {exemplos}.'
    )


def montar_preview_importacao_minuta(arquivo, usuario):
    logger.info(
        'DEBUG MINUTA INICIO: arquivo_recebido=%s user_id=%s',
        getattr(arquivo, 'name', ''),
        getattr(usuario, 'id', None),
    )
    try:
        workbook = load_workbook(arquivo, read_only=True, data_only=True)
    except Exception as exc:
        raise MinutaImportacaoError(f'Falha ao ler a planilha: {exc}') from exc
    logger.info('DEBUG MINUTA: planilha_carregada arquivo=%s', getattr(arquivo, 'name', ''))

    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    logger.info('DEBUG MINUTA: linhas_excel=%s abas=%s', len(rows), list(workbook.sheetnames))
    if not rows:
        raise MinutaImportacaoError('A planilha de romaneio está vazia.')

    meta_header_index = _localizar_linha_cabecalho(rows, ['Filial', 'Dt. Saída', 'Carga'])
    data_header_index = _localizar_linha_cabecalho(rows, ['Carregamento', 'Número Nota'])

    meta_headers = rows[meta_header_index]
    meta_values = rows[meta_header_index + 1] if meta_header_index + 1 < len(rows) else []
    meta_indices = _mapear_indices_cabecalho(meta_headers, obrigatorios={'filial', 'dtsaida', 'carga'})
    data_indices = _mapear_indices_cabecalho(rows[data_header_index], obrigatorios={'carregamento', 'numeronota'})

    data_saida = _parse_data(_obter_valor(meta_values, meta_indices, 'dtsaida'))
    veiculo = _texto_limpo(_obter_valor(meta_values, meta_indices, 'veiculo'))
    motorista = _extrair_nome_pessoa(_obter_valor(meta_values, meta_indices, 'motorista'))
    placa = _extrair_placa(veiculo)

    linhas_planilha = []
    erros = []
    for row_number, row in enumerate(rows[data_header_index + 1 :], start=data_header_index + 2):
        if not _linha_tem_conteudo(row):
            continue
        numero_nota = _texto_limpo(_obter_valor(row, data_indices, 'numeronota'))
        codigo_romaneio = _texto_limpo(_obter_valor(row, data_indices, 'carregamento'))
        if not numero_nota or not codigo_romaneio:
            erros.append(f'Linha {row_number} ignorada por ausência de Número Nota ou Carregamento.')
            continue
        linhas_planilha.append(
            {
                'codigo_romaneio': codigo_romaneio,
                'numero_nota': numero_nota,
                'sequencia_entrega': _texto_limpo(_obter_valor(row, data_indices, 'seqent')),
                'codigo_cliente': _texto_limpo(_obter_valor(row, data_indices, 'codigo')),
                'fantasia': _texto_limpo(_obter_valor(row, data_indices, 'fantasia')),
                'razao_social': _texto_limpo(_obter_valor(row, data_indices, 'razaosocial')),
                'numero_pedido': _texto_limpo(_obter_valor(row, data_indices, 'numeropedido')),
                'tipo_cobranca': _texto_limpo(_obter_valor(row, data_indices, 'tipocobranca')),
                'peso_kg': str(_parse_decimal(_obter_valor(row, data_indices, 'pesokg'))),
                'volume_m3': str(_parse_decimal(_obter_valor(row, data_indices, 'volumem3'))),
                'valor_total': str(_parse_decimal(_obter_valor(row, data_indices, 'valortotal'))),
            }
        )

    if not linhas_planilha:
        raise MinutaImportacaoError('Nenhuma linha válida foi encontrada na planilha de romaneio.')
    logger.info('DEBUG MINUTA: romaneios_encontrados=%s linhas_validas=%s', len({linha['codigo_romaneio'] for linha in linhas_planilha}), len(linhas_planilha))

    numeros_notas = {linha['numero_nota'] for linha in linhas_planilha}
    nfs = list(
        NotaFiscal.objects.select_related('cliente', 'rota')
        .filter(numero__in=numeros_notas)
        .order_by('numero', '-data_emissao', '-id')
    )
    nf_por_numero = {}
    for nf in nfs:
        nf_por_numero.setdefault(nf.numero, nf)
    logger.info('DEBUG MINUTA: nfs_encontradas=%s', len(nf_por_numero))

    entradas = list(
        EntradaNF.objects.filter(numero_nf__in=numeros_notas)
        .order_by('numero_nf', '-data_importacao', '-id')
    )
    entrada_por_numero = {}
    for entrada in entradas:
        entrada_por_numero.setdefault(entrada.numero_nf, entrada)
    xml_por_numero = _carregar_xml_entrada_por_numero(
        {
            numero_nota: entrada_por_numero[numero_nota]
            for numero_nota in numeros_notas
            if numero_nota not in nf_por_numero and numero_nota in entrada_por_numero
        },
        usuario,
    )
    logger.info('DEBUG MINUTA: xmls_encontrados=%s', len(xml_por_numero))

    diagnostico_tabelas = _diagnostico_tabelas_minuta()
    if diagnostico_tabelas['resultado_validacao']:
        itens_existentes = list(
            MinutaRomaneioItem.objects.select_related('romaneio__usuario_importacao')
            .filter(numero_nota__in=numeros_notas)
            .order_by('numero_nota', '-romaneio__data_saida', '-created_at')
        )
    else:
        itens_existentes = []
        mensagem_estrutura = _mensagem_erro_estrutura_minuta(diagnostico_tabelas)
        logger.error(mensagem_estrutura)
        erros.append(f'{mensagem_estrutura}. O preview será exibido sem validar duplicidades históricas.')
    itens_por_nota = defaultdict(list)
    for item in itens_existentes:
        itens_por_nota[item.numero_nota].append(item)

    preview_linhas = []
    romaneios = set()
    duplicados = 0
    bloqueadas = 0
    for linha in linhas_planilha:
        romaneios.add(linha['codigo_romaneio'])
        xml_info = xml_por_numero.get(linha['numero_nota'], {})
        referencia_nf = _resolver_status_minuta_nf(
            nf_por_numero.get(linha['numero_nota']),
            entrada_por_numero.get(linha['numero_nota']),
            xml_info,
        )
        nf = referencia_nf['nf']
        duplicidade = None
        for item_existente in itens_por_nota.get(linha['numero_nota'], []):
            if item_existente.romaneio.codigo_romaneio != linha['codigo_romaneio']:
                duplicidade = item_existente
                break

        status = referencia_nf['status']
        importavel = referencia_nf['importavel']
        if duplicidade is not None:
            status = 'DUPLI'
            duplicados += 1
        elif not importavel:
            bloqueadas += 1

        preview_linha = {
            **linha,
            'data_saida': _serializar_data(data_saida),
            'filial': _texto_limpo(_obter_valor(meta_values, meta_indices, 'filial')),
            'destino': _texto_limpo(_obter_valor(meta_values, meta_indices, 'destino')),
            'km': _texto_limpo(_obter_valor(meta_values, meta_indices, 'km')),
            'rotas': _texto_limpo(_obter_valor(meta_values, meta_indices, 'rotas')),
            'quantidade_pedidos': _parse_int(_obter_valor(meta_values, meta_indices, 'qtdpedidos')),
            'quantidade_clientes': _parse_int(_obter_valor(meta_values, meta_indices, 'qtdclientes')),
            'veiculo': veiculo,
            'placa': placa,
            'motorista': motorista,
            'ajudante_1': _extrair_nome_pessoa(_obter_valor(meta_values, meta_indices, 'ajudante1')),
            'ajudante_2': _extrair_nome_pessoa(_obter_valor(meta_values, meta_indices, 'ajudante2')),
            'ajudante_3': _extrair_nome_pessoa(_obter_valor(meta_values, meta_indices, 'ajudante3')),
            'numero_box': _texto_limpo(_obter_valor(meta_values, meta_indices, 'nbox')),
            'transportadora': _texto_limpo(_obter_valor(meta_values, meta_indices, 'transportadora')),
            'arquivo_nome': getattr(arquivo, 'name', ''),
            'nf_id': nf.id if nf else None,
            'cliente_nf': nf.cliente.nome if nf and nf.cliente_id else xml_info.get('cliente_nome', ''),
            'bairro_nf': _bairro_nf(nf) or xml_info.get('bairro', ''),
            'status': status,
            'importavel': importavel,
            'motivo_bloqueio': referencia_nf['motivo'],
            'duplicado': duplicidade is not None,
            'duplicidade_romaneio_codigo': duplicidade.romaneio.codigo_romaneio if duplicidade else '',
            'duplicidade_data_saida': _serializar_data(duplicidade.romaneio.data_saida) if duplicidade else '',
            'duplicidade_placa': duplicidade.romaneio.placa if duplicidade else '',
            'duplicidade_motorista': duplicidade.romaneio.motorista if duplicidade else '',
            'duplicidade_usuario': (
                getattr(duplicidade.romaneio.usuario_importacao, 'nome', '')
                or getattr(duplicidade.romaneio.usuario_importacao, 'username', '')
            )
            if duplicidade
            else '',
            'usuario_responsavel': getattr(usuario, 'nome', '') or getattr(usuario, 'username', ''),
        }
        preview_linhas.append(preview_linha)

    preview_linhas.sort(key=lambda linha: (linha['codigo_romaneio'], linha['sequencia_entrega'] or linha['numero_nota']))
    return {
        'meta': {
            'filial': _texto_limpo(_obter_valor(meta_values, meta_indices, 'filial')),
            'data_saida': _serializar_data(data_saida),
            'cargas': _texto_limpo(_obter_valor(meta_values, meta_indices, 'carga')),
            'destino': _texto_limpo(_obter_valor(meta_values, meta_indices, 'destino')),
            'km': _texto_limpo(_obter_valor(meta_values, meta_indices, 'km')),
            'rotas': _texto_limpo(_obter_valor(meta_values, meta_indices, 'rotas')),
            'quantidade_pedidos': _parse_int(_obter_valor(meta_values, meta_indices, 'qtdpedidos')),
            'quantidade_clientes': _parse_int(_obter_valor(meta_values, meta_indices, 'qtdclientes')),
            'veiculo': veiculo,
            'placa': placa,
            'motorista': motorista,
            'ajudante_1': _extrair_nome_pessoa(_obter_valor(meta_values, meta_indices, 'ajudante1')),
            'ajudante_2': _extrair_nome_pessoa(_obter_valor(meta_values, meta_indices, 'ajudante2')),
            'ajudante_3': _extrair_nome_pessoa(_obter_valor(meta_values, meta_indices, 'ajudante3')),
            'numero_box': _texto_limpo(_obter_valor(meta_values, meta_indices, 'nbox')),
            'transportadora': _texto_limpo(_obter_valor(meta_values, meta_indices, 'transportadora')),
            'arquivo_nome': getattr(arquivo, 'name', ''),
        },
        'linhas': preview_linhas,
        'erros': erros,
        'resumo': {
            'romaneios': len(romaneios),
            'itens': len(preview_linhas),
            'erros': len(erros),
            'duplicados': duplicados,
            'bloqueadas': bloqueadas,
            'atualizacao': 'Manual',
        },
    }


def confirmar_importacao_minuta(preview, usuario, validar_restricoes=True):
    linhas = preview.get('linhas', [])
    logger.info('DEBUG MINUTA: confirmacao_inicio linhas=%s user_id=%s', len(linhas), getattr(usuario, 'id', None))
    if not linhas:
        raise MinutaImportacaoError('Nenhuma prévia de importação da minuta foi encontrada para confirmação.')
    diagnostico_tabelas = _diagnostico_tabelas_minuta()
    if not diagnostico_tabelas['resultado_validacao']:
        logger.warning(
            'DEBUG MINUTA: validacao_estrutura_nao_bloqueante schema=%s alias=%s tabelas_encontradas=%s tabelas_faltantes=%s erro=%s',
            diagnostico_tabelas.get('schema_detectado'),
            diagnostico_tabelas.get('alias'),
            diagnostico_tabelas.get('tabelas_encontradas'),
            diagnostico_tabelas.get('tabelas_faltantes'),
            diagnostico_tabelas.get('erro'),
        )
    if validar_restricoes:
        _validar_preview_minuta_confirmavel(preview)

    data_saida = _parse_data(preview.get('meta', {}).get('data_saida'))
    linhas_por_romaneio = defaultdict(list)
    for linha in linhas:
        linhas_por_romaneio[linha['codigo_romaneio']].append(linha)

    romaneios_processados = 0
    itens_processados = 0
    duplicados = 0
    logger.info('DEBUG MINUTA: criando_minuta romaneios=%s', len(linhas_por_romaneio))
    with transaction.atomic():
        codigos_romaneio = list(linhas_por_romaneio.keys())
        lote_importacao = uuid.uuid4()
        MinutaRomaneio.objects.filter(codigo_romaneio__in=codigos_romaneio).delete()

        for codigo_romaneio, linhas_romaneio in linhas_por_romaneio.items():
            logger.info('DEBUG MINUTA: criando_romaneio codigo=%s itens=%s', codigo_romaneio, len(linhas_romaneio))
            romaneio = MinutaRomaneio.objects.create(
                codigo_romaneio=codigo_romaneio,
                importacao_lote=lote_importacao,
                data_saida=data_saida,
                filial=preview['meta'].get('filial', ''),
                destino=preview['meta'].get('destino', ''),
                km=preview['meta'].get('km', ''),
                rotas=preview['meta'].get('rotas', ''),
                quantidade_pedidos=preview['meta'].get('quantidade_pedidos'),
                quantidade_clientes=preview['meta'].get('quantidade_clientes'),
                veiculo=preview['meta'].get('veiculo', ''),
                placa=preview['meta'].get('placa', ''),
                motorista=preview['meta'].get('motorista', ''),
                ajudante_1=preview['meta'].get('ajudante_1', ''),
                ajudante_2=preview['meta'].get('ajudante_2', ''),
                ajudante_3=preview['meta'].get('ajudante_3', ''),
                numero_box=preview['meta'].get('numero_box', ''),
                transportadora=preview['meta'].get('transportadora', ''),
                arquivo_nome=preview['meta'].get('arquivo_nome', ''),
                usuario_importacao=usuario,
            )
            romaneios_processados += 1

            itens_novos = []
            for linha in linhas_romaneio:
                razao_social = linha.get('razao_social', '')
                if not linha.get('nf_id') and linha.get('cliente_nf'):
                    razao_social = linha['cliente_nf']
                itens_novos.append(
                    MinutaRomaneioItem(
                        romaneio=romaneio,
                        nf_id=linha.get('nf_id'),
                        numero_nota=linha['numero_nota'],
                        sequencia_entrega=linha.get('sequencia_entrega', ''),
                        codigo_cliente=linha.get('codigo_cliente', ''),
                        fantasia=linha.get('fantasia', ''),
                        razao_social=razao_social,
                        bairro=linha.get('bairro_nf', '') or linha.get('bairro', ''),
                        numero_pedido=linha.get('numero_pedido', ''),
                        tipo_cobranca=linha.get('tipo_cobranca', ''),
                        peso_kg=_parse_decimal(linha.get('peso_kg')),
                        volume_m3=_parse_decimal(linha.get('volume_m3')),
                        valor_total=_parse_decimal(linha.get('valor_total')),
                        status=linha.get('status', 'PENDENTE'),
                        duplicado=bool(linha.get('duplicado')),
                        duplicidade_romaneio_codigo=linha.get('duplicidade_romaneio_codigo', ''),
                        duplicidade_data_saida=_parse_data(linha.get('duplicidade_data_saida')),
                        duplicidade_placa=linha.get('duplicidade_placa', ''),
                        duplicidade_motorista=linha.get('duplicidade_motorista', ''),
                        duplicidade_usuario=linha.get('duplicidade_usuario', ''),
                    )
                )
                duplicados += 1 if linha.get('duplicado') else 0
                itens_processados += 1

            if itens_novos:
                logger.info('DEBUG MINUTA: criando_itens romaneio=%s quantidade=%s', codigo_romaneio, len(itens_novos))
                MinutaRomaneioItem.objects.bulk_create(itens_novos)

    logger.info('DEBUG MINUTA: finalizando_importacao romaneios=%s itens=%s duplicados=%s', romaneios_processados, itens_processados, duplicados)

    return {
        'romaneios': romaneios_processados,
        'itens': itens_processados,
        'duplicados': duplicados,
    }


def _obter_lote_minuta_ativo():
    try:
        return (
            MinutaRomaneio.objects.order_by('-created_at', '-id')
            .values_list('importacao_lote', flat=True)
            .first()
        )
    except (ProgrammingError, OperationalError):
        logger.exception('ERRO REAL MINUTA: falha ao obter lote ativo da minuta.')
        return None


def consultar_minuta_itens_queryset(romaneio='', status='', busca=''):
    try:
        queryset = MinutaRomaneioItem.objects.select_related('romaneio', 'nf', 'nf__rota', 'romaneio__usuario_importacao').all()
        lote_ativo = _obter_lote_minuta_ativo()
        if lote_ativo:
            queryset = queryset.filter(romaneio__importacao_lote=lote_ativo)
        else:
            queryset = queryset.none()
        if romaneio:
            queryset = queryset.filter(romaneio__codigo_romaneio__icontains=romaneio)
        if status:
            queryset = queryset.filter(status=status)
        if busca:
            queryset = queryset.filter(
                Q(numero_nota__icontains=busca)
                | Q(fantasia__icontains=busca)
                | Q(razao_social__icontains=busca)
                | Q(romaneio__placa__icontains=busca)
                | Q(romaneio__motorista__icontains=busca)
                | Q(romaneio__codigo_romaneio__icontains=busca)
                | Q(status__icontains=busca)
            )

        return queryset.order_by('-romaneio__data_saida', 'romaneio__codigo_romaneio', 'numero_nota')
    except (ProgrammingError, OperationalError):
        logger.exception('ERRO REAL MINUTA: falha ao consultar itens da minuta.')
        return MinutaRomaneioItem.objects.none()


def listar_minuta_itens(romaneio='', status='', busca=''):
    queryset = consultar_minuta_itens_queryset(romaneio=romaneio, status=status, busca=busca)
    linhas = [
        {
            'romaneio': item.romaneio.codigo_romaneio,
            'nf_id': item.nf_id,
            'numero_nota': item.numero_nota,
            'data_saida': item.romaneio.data_saida.strftime('%d/%m/%Y') if item.romaneio.data_saida else '-',
            'motorista': item.romaneio.motorista or '-',
            'placa': item.romaneio.placa or '-',
            'cliente': item.razao_social or item.fantasia or (item.nf.cliente.nome if item.nf_id else '-'),
            'bairro': item.bairro or _bairro_nf(item.nf) or '-',
            'status': item.status,
            'duplicado': item.duplicado,
            'duplicidade_texto': (
                f"NF vinculada ao romaneio {item.duplicidade_romaneio_codigo} em {item.duplicidade_data_saida.strftime('%d/%m/%Y') if item.duplicidade_data_saida else '-'}"
                if item.duplicado
                else ''
            ),
            'peso_kg': f'{item.peso_kg:.3f}'.replace('.', ','),
            'valor_total': f'{item.valor_total:.2f}'.replace('.', ','),
        }
        for item in queryset
    ]

    inconsistencias = get_minuta_inconsistencias(linhas)

    resumo = {
        'romaneios': len({linha['romaneio'] for linha in linhas}),
        'itens': len(linhas),
        'erros': 0,
        'duplicados': sum(1 for linha in linhas if linha['duplicado']),
        'bloqueadas': sum(1 for linha in linhas if linha['status'] in {'NF CANCELADA', 'NF DENEGADA', 'NF INCONSISTENTE', 'NF BLOQUEADA', 'NF INATIVA', 'NF COM PROBLEMA', 'XML INVALIDO', 'NF NÃO LOCALIZADA'}),
        'nfs_nao_encontradas': inconsistencias['nfs_nao_encontradas'],
        'xml_erros': inconsistencias['xml_erros'],
        'inconsistencias': inconsistencias['inconsistencias'],
        'atualizacao': 'Manual',
    }
    return linhas, resumo